import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# تنظیمات مسیرها
input_folder = ".."
output_file = "src_programing/merged.xlsx"
excel_files = ["1403.xlsx", "Digikala.xlsx", "others.xlsx"]

# تعریف ستون‌های خروجی با ترتیب مورد نظر
OUTPUT_COLUMNS = [
    'منبع فایل',
    'نام شیت',
    'تاریخ استخراج',
    'طاهر',
    'مجتبی',
    'محمد',
    'تنخواه',
    'سود',
    'مایه',
    'طرف حساب'
]

# ستون‌های عددی که نیاز به محاسبه جمع دارند
NUMERIC_COLUMNS = ['طاهر', 'مجتبی', 'محمد', 'تنخواه', 'سود', 'مایه']

# استایل‌ها
HEADER_STYLE = {
    'font': Font(bold=True, color='FFFFFF'),
    'fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
    'border': Border(bottom=Side(border_style='thin', color='000000'))
}

TOTAL_STYLE = {
    'font': Font(bold=True, color='000000'),
    'fill': PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid'),  # رنگ طلایی
    'border': Border(top=Side(border_style='double', color='000000'))
}


# در تابع apply_styles



# در تابع process_files (قبل از ذخیره فایل)

# ببببببببببببببببببببببببببب

def extract_fff_rows(df, file_name, sheet_name):
    """استخراج ردیف‌های حاوی 'fff' و تبدیل به ساختار مورد نظر"""
    fff_rows = df[df.iloc[:, 0].astype(str).str.strip() == 'fff']

    results = []
    for _, row in fff_rows.iterrows():
        extracted_data = {
            'منبع فایل': file_name,
            'نام شیت': sheet_name,
            'تاریخ استخراج': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'طاهر': '',
            'مجتبی': '',
            'محمد': '',
            'تنخواه': '',
            'سود': '',
            'مایه': '',
            'طرف حساب': ''
        }

        # استخراج مقادیر از ستون‌های مختلف
        for col in df.columns:
            col_value = row[col] if pd.notna(row[col]) else ''

            # تشخیص خودکار ستون‌ها بر اساس محتوا
            if 'طاهر' in str(col):
                extracted_data['طاهر'] = col_value
            elif 'مجتبی' in str(col):
                extracted_data['مجتبی'] = col_value
            elif 'محمد' in str(col):
                extracted_data['محمد'] = col_value
            elif 'تنخواه' in str(col):
                extracted_data['تنخواه'] = col_value
            elif 'سود' in str(col):
                extracted_data['سود'] = col_value
            elif 'مایه' in str(col):
                extracted_data['مایه'] = col_value
            elif 'طرف حساب' in str(col):
                extracted_data['طرف حساب'] = col_value

        results.append(extracted_data)

    return results


def add_total_row(df):
    """اضافه کردن ردیف جمع به دیتافریم"""
    total_row = {'منبع فایل': 'جمع کل', 'نام شیت': '', 'تاریخ استخراج': ''}

    for col in NUMERIC_COLUMNS:
        if col in df.columns:
            # تبدیل مقادیر به عدد و محاسبه جمع
            numeric_values = pd.to_numeric(df[col], errors='coerce')
            total = numeric_values.sum()
            total_row[col] = total
        else:
            total_row[col] = ''

    # اضافه کردن ردیف جمع به انتهای دیتافریم
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    return df


def apply_styles(worksheet, df):
    """اعمال استایل به هدر و ردیف جمع"""
    # استایل هدر (آبی)
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = HEADER_STYLE['font']
        cell.fill = HEADER_STYLE['fill']
        cell.border = HEADER_STYLE['border']
        worksheet.column_dimensions[get_column_letter(col_num)].width = max(len(str(column_title)), 10) + 2

    # استایل ردیف جمع (طلایی)
    last_row = worksheet.max_row
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=last_row, column=col_num)
        cell.font = TOTAL_STYLE['font']
        cell.fill = TOTAL_STYLE['fill']
        cell.border = TOTAL_STYLE['border']

        if col_num > 3 and col_num <= 3 + len(NUMERIC_COLUMNS):
            column_letter = get_column_letter(col_num)
            cell.value = f"=SUM({column_letter}2:{column_letter}{last_row - 1})"


def process_files():
    """پردازش فایل‌ها و استخراج ردیف‌های fff"""
    output_path = os.path.join(input_folder, output_file)
    all_results = []

    for file in excel_files:
        file_path = os.path.join(input_folder, file)

        if not os.path.exists(file_path):
            print(f"⚠️ فایل {file} یافت نشد، رد شد.")
            continue

        try:
            xls = pd.ExcelFile(file_path)

            for sheet_name in xls.sheet_names:
                try:
                    df = pd.read_excel(xls, sheet_name=sheet_name)

                    if df.empty:
                        print(f"⚠️ شیت '{sheet_name}' در فایل {file} خالی است، رد شد.")
                        continue

                    sheet_results = extract_fff_rows(df, file, sheet_name)
                    all_results.extend(sheet_results)

                except Exception as e:
                    print(f"❌ خطا در پردازش شیت '{sheet_name}' از فایل {file}: {str(e)}")
                    continue

        except Exception as e:
            print(f"❌ خطا در خواندن فایل {file}: {str(e)}")
            continue

    if all_results:
        result_df = pd.DataFrame(all_results)

        # اطمینان از وجود تمام ستون‌های مورد نیاز
        for col in OUTPUT_COLUMNS:
            if col not in result_df.columns:
                result_df[col] = ''

        # مرتب کردن ستون‌ها
        result_df = result_df[OUTPUT_COLUMNS]

        # اضافه کردن ردیف جمع
        result_df = add_total_row(result_df)

        # جابجایی ستون طرف حساب به آخر
        result_df = result_df[[c for c in result_df.columns if c != 'طرف حساب'] + ['طرف حساب']]


        # ذخیره فایل با استایل‌ها
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='نتایج', index=False)

            worksheet = writer.sheets['نتایج']
            worksheet.sheet_view.rightToLeft = True

            # اعمال استایل‌ها
            apply_styles(worksheet, result_df)

        print(f"✅ عملیات با موفقیت انجام شد. {len(result_df) - 1} ردیف به فایل خروجی اضافه شد.")
        print(output_file)
    else:
        print("⚠️ هیچ ردیف fff یافت نشد.")


if __name__ == "__main__":
    process_files()